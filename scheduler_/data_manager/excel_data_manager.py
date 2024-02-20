import pandas as pd
from openpyxl import load_workbook


class ExcelDataManager:
    def __init__(self, excel_file_path: str):
        self.excel_file_path = excel_file_path
        self.df = None

    def load_excel_data(self):
        """Loads data from the Excel file into a pandas DataFrame."""
        self.df = pd.read_excel(self.excel_file_path)

    def save_changes_to_excel(self):
        """Saves changes from the DataFrame's 'Status' column back to the Excel file based on 'Post ID'."""
        # Load the workbook and the specific sheet
        workbook = load_workbook(self.excel_file_path)
        sheet_name = 'Sheet1'  # Replace with your actual sheet name
        sheet = workbook[sheet_name]

        # Assuming 'Post ID' is the first column in both the Excel file and the DataFrame
        post_id_col_idx = 1  # Excel columns start from 1
        status_col_idx = sheet.max_column  # Assuming 'Status' is the last column

        # Create a dictionary of 'Post ID' to 'Status' from the DataFrame for quick look-up
        status_dict = pd.Series(self.df.Status.values, index=self.df['Post ID']).to_dict()

        # Iterate over the rows in the Excel sheet to update the 'Status' column
        for row in sheet.iter_rows(min_row=2, max_col=post_id_col_idx, max_row=sheet.max_row):
            post_id_cell = row[0]
            if post_id_cell.value in status_dict:
                # Update the 'Status' column in the sheet
                sheet.cell(row=post_id_cell.row, column=status_col_idx, value=status_dict[post_id_cell.value])

        # Save the workbook
        workbook.save(self.excel_file_path)
