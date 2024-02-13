# scheduler/scheduler.py

import pandas as pd
from apscheduler.schedulers.background import BackgroundScheduler
from openpyxl.reader.excel import load_workbook

from post_manager.bot_core.posts import SocialMediaPost
from logger_config import logger, console
from rich.table import Table
from datetime import datetime
import importlib
import os


def display_dataframe_as_table(dataframe):
    table = Table(show_header=True, header_style="bold magenta")
    for column in dataframe.columns:
        table.add_column(column)

    for _, row in dataframe.iterrows():
        table.add_row(*[str(row[column]) for column in dataframe.columns])

    console.print(table)


class ExcelScheduler:
    def __init__(self, excel_file):
        """
        Initializes the ExcelScheduler with the specified Excel file.
        It sets up a background scheduler for running scheduled jobs and
        loads the available platform-specific posting classes dynamically.

        :param excel_file: The path to the Excel file that contains the posting schedule.
        """
        self.df = None
        self.excel_file = excel_file
        self.scheduler = BackgroundScheduler()
        # self.platform_post_classes = self.load_platform_post_classes()

    def start(self):
        """
        Starts the scheduler to begin running scheduled jobs. This method
        sets up the daily data loading and post scheduling, then starts the scheduler.
        """
        # Set up daily scheduling and execute today's scheduling upon starting
        self.schedule_daily_loading()
        self.load_data_and_schedule()
        # Start the scheduler
        self.scheduler.start()

    def stop(self):
        """
        Shuts down the scheduler and all its running jobs. This should be called
        before exiting the application to ensure a clean shutdown.
        """
        self.scheduler.shutdown()

    def load_excel_data(self):
        """
        Loads data from the Excel file into a pandas DataFrame. This method is typically
        called to refresh the data before scheduling new posts.
        """
        # Read the Excel file
        self.df = pd.read_excel(self.excel_file)

    def load_data_and_schedule(self):
        """
        Loads data from the Excel file using 'load_excel_data' and schedules posts for
        the current day using 'schedule_today_posts'.
        """
        self.load_excel_data()
        self.schedule_today_posts()

    def schedule_daily_loading(self):
        """
        Sets up a daily job in the scheduler to load data from the Excel file and
        schedule posts every day at 01:00 AM.
        """
        # Schedule to read Excel file every day at 01:00 AM
        self.scheduler.add_job(self.load_data_and_schedule, 'cron', hour=1, minute=0)
        # TODO: 24. 1. 2024: Add schedule for update excel file

    def schedule_today_posts(self):
        """
        Schedules the posts that are due for the current day. It filters the DataFrame
        for today's posts and schedules each one based on the specified time.
        """
        # Get current date and
        current_date = datetime.now().date()
        # Filter posts scheduled for the next day
        today_posts = self.df[(self.df['Scheduled Time'].dt.date == current_date) & (self.df['Status'] == 'Scheduled')]

        # Display the filtered DataFrame as a table
        display_dataframe_as_table(today_posts)

        # Schedule posts based on the data from Excel file
        for _, row in today_posts.iterrows():
            post_time = row['Scheduled Time']
            # Pass 'row' as a default argument to the lambda function
            self.scheduler.add_job(lambda current_row=row: self.post_content(current_row),
                                   'date',
                                   run_date=post_time)

    def load_platform_post_classes(self):
        """
        Dynamically loads and creates a dictionary mapping of platform names to their
        respective post classes from the 'post_manager/bots' directory. This allows for
        the dynamic instantiation of post classes when scheduling content.

        :return: A dictionary mapping platform names to their respective post classes.
        """
        platform_classes = {}
        bots_dir = 'post_manager/bots'  # Directory where bot modules are located
        for file in os.listdir(bots_dir):
            if file.endswith('.py') and file != '__init__.py':
                module_name = file[:-3]  # Strip '.py'
                module = importlib.import_module(f"post_manager.bots.{module_name}")
                for attr_name in dir(module):
                    attr = getattr(module, attr_name)
                    if isinstance(attr, type) and issubclass(attr, SocialMediaPost) and attr is not SocialMediaPost:
                        platform_classes[module_name] = attr
        return platform_classes

    def load_bot(self, platform_name):
        """
        Dynamically loads a bot class based on the given platform name from the
        'post_manager/bots' directory and instantiates it.

        :param platform_name: The name of the platform for which to load the bot class.
        :return: An instance of the bot class if successful, or None if an error occurs.
        """
        try:
            # Capitalize the first letter to match the class name convention
            class_name = platform_name.capitalize() + 'Bot'
            module = importlib.import_module(f"post_manager.bots.{platform_name}")

            logger.info(f"Loaded class {class_name} from post_manager.bots.{platform_name}")

            bot_class = getattr(module, class_name)
            return bot_class(self.excel_file)  # Pass the Excel file name to the bot constructor
        except (ImportError, AttributeError) as e:
            logger.error(f"Error loading bot for platform {platform_name}: {e}")
            return None

    def post_content(self, row):
        """
        Invokes the posting process for the given row of data from the DataFrame.
        It uses the appropriate bot for the platform specified in the row.

        :param row: A pandas Series object representing the data for a single post.
        """
        platform = row['Platform'].lower()
        bot = self.load_bot(platform)

        if bot:
            # Dynamically load the post class from the module, not the bot instance
            post_module = importlib.import_module(f"post_manager.bots.{platform}")
            post_class_name = platform.capitalize() + 'Post'
            post_class = getattr(post_module, post_class_name)

            # Create an instance of the post class
            post = post_class(post_id=row['Post ID'],
                              content=row['Content'],
                              image_path=row['Image Path'],
                              hashtags=row['Hashtags'])

            # Call the post method of the bot with the post instance
            bot.post(post)

            # TODO: 4. 2. 2024: Display post in table
            # TODO: Marek: Change dataframe if is successfully posted
            # TODO: Marek: Call save_dataframe_changes_to_excel
        else:
            logger.error(f"No bot available for platform: {platform}")

    def save_dataframe_changes_to_excel(self, excel_file_path):
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active

        # Assuming the first row is the header and 'Post ID' is in the first column
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        post_id_col_idx = header.index('Post ID') + 1  # +1 because Excel is 1-indexed
        status_col_idx = header.index('Status') + 1
        last_updated_col_idx = header.index('Last Updated') + 1 if 'Last Updated' in header else None

        # Iterate over the DataFrame rows
        for _, row in self.df.iterrows():
            # Find the Excel row corresponding to the DataFrame row
            for excel_row in sheet.iter_rows(min_row=2):  # Skip header row
                excel_post_id = excel_row[post_id_col_idx - 1].value  # -1 because Excel is 1-indexed
                if excel_post_id == row['Post ID']:
                    # Update the 'Status' column in Excel
                    excel_row[status_col_idx - 1].value = row['Status']
                    # If there's a 'Last Updated' column, update it too
                    if last_updated_col_idx:
                        excel_row[last_updated_col_idx - 1].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    break

        workbook.save(excel_file_path)
        workbook.close()

    # TODO: 2. 2. 2024: All functions for update excel file must be change for new format
    def update_excel_file_from_logs(self, only_today=False):
        """
        Updates the status in the Excel file for each post based on the corresponding
        log entries. It has an option to filter for today's entries only.

        :param only_today: A boolean indicating whether to update only today's entries.
        """
        log_entries = self.read_log_entries(only_today)
        workbook = load_workbook(self.excel_file)
        sheet = workbook.active

        # TODO: 2. 2. 2024: Index can be dynamic and not static
        status_column_index = 7  # Assuming 'Status' is the third column

        logger.debug(f"Log entries: {log_entries}")  # Debug print

        for entry in log_entries:
            for row in sheet.iter_rows():
                # Assuming 'Post ID' is in the first column
                post_id_cell = row[0]
                if post_id_cell.value == entry['post_id']:
                    status_cell = row[status_column_index - 1]  # Subtract 1 because Python uses 0-based indexing
                    status_cell.value = entry['status']
                    logger.debug(f"Updating Post ID {entry['post_id']} status to {entry['status']}")  # Debug print
                    break

        workbook.save(self.excel_file)
        workbook.close()

    def read_log_entries(self, only_today):
        """
        Reads log entries from log files located in the 'post_manager/logs' directory.
        It can optionally include only today's log entries based on the 'only_today' parameter.

        :param only_today: A boolean indicating whether to include only today's log entries.
        :return: A list of log entry dictionaries with 'post_id' and 'status' keys.
        """
        logs_dir = 'post_manager/logs'
        log_entries = []
        logger.debug(f"Reading log entries from {os.listdir(logs_dir)}")
        for log_file in os.listdir(logs_dir):
            if log_file.endswith('_posts.log'):
                logger.debug(f"Log file name: {log_file}")
                full_path = os.path.join(logs_dir, log_file)
                with open(full_path, 'r') as file:
                    for line in file:
                        if only_today and not self.is_entry_for_today(line):
                            continue
                        entry = self.parse_log_line(line)
                        if entry:
                            log_entries.append(entry)
        return log_entries

    def is_entry_for_today(self, line):
        """
        Checks if a log entry line corresponds to the current date.

        :param line: A string representing a line from a log file.
        :return: A boolean indicating whether the line is an entry for today.
        """
        log_date_str = line.split()[0]
        try:
            log_date = datetime.strptime(log_date_str, '%Y-%m-%d').date()
            is_today = log_date == datetime.now().date()
            logger.debug(
                f"Is entry for today: {is_today}, Line date: {log_date}, Today: {datetime.now().date()}")  # Debug print
            return is_today
        except ValueError as e:
            logger.debug(f"Error parsing date: {e}")  # Debug print
            return False

    def parse_log_line(self, line):
        """
        Parses a single line from the log file to extract the post ID and status.
        The expected format of the log line is:
        'YYYY-MM-DD HH:MM:SS,SSS - logger_name - INFO - Post ID X: Status message'

        :param line: A string representing a line from a log file.
        :return: A dictionary with 'post_id' and 'status' extracted from the line, or None if parsing fails.
        """
        # TODO: 2. 2. 2024: Main function for read line and separate data. Change to new format or read dynamically
        try:
            # First, split the line by ' - ' which separates the date, logs name, log level, and message
            parts = line.split(' - ')
            # Check if the split parts list has at least 4 elements (date, logs, level, message)
            if len(parts) >= 4:
                # The message part containing 'Post ID' and the status message is the fourth element
                message = parts[3]
                # Now we split the message by ' ', expecting a format like 'Post ID 10: Successfully...'
                message_parts = message.split(' ')
                # Check if the first two elements in message_parts are 'Post' and 'ID'
                if message_parts[:2] == ['Post', 'ID']:
                    # Extract the post ID, remove any trailing colons, and convert to an integer
                    post_id = int(message_parts[2].rstrip(':').strip())
                    # The rest of the message is the status, joined back into a string
                    status = ' '.join(message_parts[3:])
                    # Return a dictionary with the post ID and status
                    return {'post_id': post_id, 'status': status}
            else:
                # If the line doesn't match the expected format, log it for debugging
                logger.debug(f"Could not parse log line: {line}")
        except Exception as e:
            # If any error occurs during parsing, log the error and the problematic line
            logger.error(f"Error parsing log line: {line}. Error: {e}")
        # Return None if the line wasn't parsed successfully
        return None
