from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


# Ensure the class documentation is descriptive
class ExcelDataManager:
    def __init__(self,
                 excel_file_path: Path,
                 sheet_name: str = 'Sheet1',
                 id_column: str = 'Post ID',
                 status_column: str = 'Status',
                 date_column: str = 'Scheduled Time'):
        """
        Initializes the ExcelDataManager with the specified Excel file path and configuration.

        :param excel_file_path: Path to the Excel file.
        :param sheet_name: Name of the sheet to be managed.
        :param id_column: Column header for unique identifiers.
        :param status_column: Column header for status entries.
        :param date_column: Column header for the scheduled date and time of posts.
        """
        self.excel_file_path = excel_file_path
        self.sheet_name = sheet_name
        self.id_column = id_column
        self.status_column = status_column
        self.date_column = date_column
        self.df = None
        self.load_excel_data()  # Load the Excel data upon initialization.

    def load_excel_data(self):
        """Loads the entire Excel sheet into a pandas DataFrame."""
        try:
            self.df = pd.read_excel(self.excel_file_path, sheet_name=self.sheet_name)

            # Ensure 'User Name' column exists, fill empty values with 'default'
            if 'User Name' in self.df.columns:
                self.df['User Name'] = self.df['User Name'].fillna('default')
            else:
                # If 'User Name' column doesn't exist, add it with 'default' for all rows
                self.df['User Name'] = 'default'

        except Exception as e:
            # Consider logging the error instead of printing to handle it properly.
            print(f"Error loading Excel data: {e}")

    def load_current_date_posts(self):
        """
        Loads posts from the Excel data that are scheduled for the current date into a DataFrame.

        :return: A DataFrame containing posts scheduled for the current date.
        """
        try:
            if self.df is not None:
                current_date = datetime.now().date()
                return self.df[self.df[self.date_column].dt.date == current_date]
            else:
                return pd.DataFrame()
        except Exception as e:
            # Consider logging the error instead of printing to handle it properly.
            print(f"Error loading current date posts: {e}")
            return pd.DataFrame()

    def update_main_dataframe(self, updated_subset_df):
        """
        Updates the main DataFrame with the changes from a subset DataFrame.

        :param updated_subset_df: The subset DataFrame with updates to be merged back into the main DataFrame.
        """
        if updated_subset_df.empty or self.df is None:
            return  # Optionally, raise an error or handle this case as appropriate for your application

        # Ensuring 'Post ID' is set as the index for proper merging
        self.df.set_index(self.id_column, inplace=True, drop=False)
        updated_subset_df.set_index(self.id_column, inplace=True, drop=False)

        # Update the main DataFrame with the subset
        self.df.update(updated_subset_df)

        # Reset the index to revert to the original structure
        self.df.reset_index(drop=True, inplace=True)

    def save_changes_to_excel(self):
        """
        Saves the changes from the DataFrame's 'Status' column back to the Excel file based on 'Post ID'.
        It updates the 'Status' column in the Excel sheet for each matching 'Post ID'.
        """
        try:
            workbook = load_workbook(self.excel_file_path)
            sheet = workbook[self.sheet_name]
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            id_col_idx = header_row.index(self.id_column) + 1
            status_col_idx = header_row.index(self.status_column) + 1

            # Update the cells in the Status column for matching Post ID
            for row in sheet.iter_rows(min_row=2, max_col=id_col_idx, max_row=sheet.max_row):
                post_id_cell = row[0]
                cell_value = self.df.loc[self.df[self.id_column] == post_id_cell.value, self.status_column].values
                if cell_value.size > 0:
                    sheet.cell(row=post_id_cell.row, column=status_col_idx, value=cell_value[0])

            workbook.save(self.excel_file_path)
            workbook.close()
        except Exception as e:
            # Consider logging the error instead of printing to handle it properly.
            print(f"Error saving changes to Excel: {e}")
